import pandas as pd
import torch
import numpy as np
import torch.nn as nn
from torch.utils.data import Dataset, DataLoader
from torchvision import transforms
import tqdm


def read_csv(filePath):
    dframe_train_feature = pd.read_excel(filePath, sheet_name='Train Feature', header=None)
    # dframe_train_label = pd.read_excel(filePath, sheet_name='Train Label', header=None)
    # dframe_test_feature = pd.read_excel(filePath, sheet_name='Test Feature', header=None)
    # dframe_test_label = pd.read_excel(filePath, sheet_name='Test Label', header=None)

    # cat1 = dframe_train_feature.head()
    # cat2 = dframe_train_label.head()
    # cat3 = dframe_test_feature.head()
    # cat4 = dframe_test_label.head()
    # print(cat1, '\n', cat2, '\n', cat3, '\n', cat4)
    train_feature = dframe_train_feature.values
    print(train_feature.shape)


class UpspDataset(Dataset):
    def __init__(self, filePath, train=True, transform=None):
        if train:
            self.dframe_feature = pd.read_excel(filePath, sheet_name='Train Feature', header=None)
            self.dframe_label = pd.read_excel(filePath, sheet_name='Train Label', header=None)
        else:
            self.dframe_feature = pd.read_excel(filePath, sheet_name='Test Feature', header=None)
            self.dframe_label = pd.read_excel(filePath, sheet_name='Test Label', header=None)

        self.features = self.dframe_feature.values.reshape(-1, 1, 16, 16)
        self.train = train

    def __getitem__(self, index):
        feature = self.features[index]
        if self.train:
            # print('check')
            labels = self.dframe_label.values
            label = labels[index]
            label = self.transform2onehot(label)
        else:
            return torch.from_numpy(feature)
        return torch.from_numpy(feature), label

    def __len__(self):
        return len(self.dframe_feature.values)

    def transform2onehot(self, digit):
        one_hot = torch.zeros(1, 10)
        one_hot[0][digit - 1] = 1
        return one_hot


class Model(nn.Module):
    def __init__(self, ):
        super(Model, self).__init__()

        self.layer1 = nn.Sequential(
            nn.Conv2d(1, 16, kernel_size=3),  # (b, 16, 14, 14)
            nn.BatchNorm2d(16),
            nn.ReLU(inplace=True)
        )

        self.layer2 = nn.Sequential(
            nn.Conv2d(16, 32, kernel_size=3),  # (b, 32, 12, 12)
            nn.BatchNorm2d(32),
            nn.ReLU(inplace=True),  # 激活函数
            nn.MaxPool2d(kernel_size=2, stride=2)  # (b, 32, 6, 6)
        )

        self.layer3 = nn.Sequential(
            nn.Conv2d(32, 64, kernel_size=3),  # (b, 64, 4, 4)
            nn.BatchNorm2d(64),
            nn.ReLU(inplace=True)
        )

        self.fc = nn.Sequential(
            nn.Linear(64 * 4 * 4, 512),
            nn.ReLU(inplace=True),
            nn.Linear(512, 128),
            nn.ReLU(inplace=True),
            nn.Linear(128, 10)
        )

    def forward(self, x):
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        x = x.view(x.size(0), -1)
        output = self.fc(x)
        return output


def train(epochs):
    dataset = UpspDataset('ups.xlsx', train=True, transform=transforms.ToTensor())
    dataloader = DataLoader(dataset, batch_size=32, shuffle=False, num_workers=3, drop_last=False)

    optimizer = torch.optim.SGD(net.parameters(), lr=0.1, momentum=0.9)
    criterion = nn.CrossEntropyLoss()

    running_loss = 0
    net.train()
    for e in range(epochs):
        batch_idx = -1
        for inputs, labels in tqdm.tqdm(dataloader):
            batch_idx += 1
            optimizer.zero_grad()
            inputs, labels = inputs.float(), labels.long()
            output = net(inputs)
            labels = labels.squeeze()
            # print(output.size(), labels.size())
            loss = criterion(output, torch.max(labels, 1)[1])
            loss.backward()

            # running_loss += loss.item()
            optimizer.step()

            if batch_idx % 50 == 0:
                print('Train Epoch: {} [{}/{} ({:.0f}%)]\tLoss: {:.6f}'.format(
                    e, batch_idx * len(inputs), len(dataset),
                       100. * batch_idx / len(dataloader), loss.item()))
    # save
    torch.save(net.state_dict(), 'model.pth')


# test
def test(length):
    net.load_state_dict(torch.load('model.pth'))
    dataset = UpspDataset('ups.xlsx', train=False)
    dataloader = DataLoader(dataset, batch_size=length, shuffle=False)
    net.eval()
    res = []
    for inputs in dataloader:
        inputs = inputs.float()
        output = net(inputs)
        digits = torch.argmax(output, 1) + 1  # 0-9 => 1-10
        res.extend(digits.tolist())

    return res


def write(digits):
    # write xlsx
    import openpyxl

    workbook = openpyxl.load_workbook('ups.xlsx')
    sheet = workbook['Test Label']
    row_max = sheet.max_row
    print(row_max)
    # sheet.insert_cols(1)

    for i in range(len(digits)):
        sheet.cell(row=i + 1, column=1, value=digits[i])
    workbook.save('ups2.xlsx')


if __name__ == '__main__':
    net = Model()
    train(30)  # 训练次数, 当前目录下没有model.pth文件时，进行训练
    res = test(length=1260)  # 使用

    # 写入
    write(res)
